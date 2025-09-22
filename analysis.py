#!/usr/bin/env python3

# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# See LICENSE for more details.
#
# Copyright: 2023 IBM
# Author: Misbah Anjum N <misanjum@linux.vnet.ibm.com>


'''
    analysis.py script is aimed to help in analysis/comparison of avocado test runs
    by generating a simple excel (.xlsx) and html (.html) files. The results.json file
    which gets created after avocado run is passed as input in command line while running
    this script and depending on the flag/options provided, the excel analysis/comparison
    sheet will be generated.

    Prerequsites:-
    pip3 install pandas[excel]
    (or)
    dnf install python3-pandas python3-numpy python3-openpyxl

    flags/options:-
    1. --new-analysis
    2. --compare-two-results

    python3 analysis.py --new-analysis <json_file>
    python3 analysis.py --compare-two-results <old_json_file> <new_json_file>

    Check README.md for more explanation
'''


import sys
import json
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


def test_analysis(data):
    '''
        This function is used to generate an excel file which contains a summary of
        avaocado based test runs. It reads results.json as input and generates a .xlsx
        file as output.
    '''

    # Extract job name from debuglog attribute
    job_name = data['debuglog'].split('/')[-2]

    # Create a DataFrame
    dataframe = pd.DataFrame(columns=['Name', 'Status', 'Fail Reason'])

    # Add rows for job name and other attributes
    dataframe.loc[0] = ['', 'Job Name', job_name]
    dataframe.loc[1] = ['', 'Fail', data['failures']]
    dataframe.loc[2] = ['', 'Error', data['errors']]
    dataframe.loc[3] = ['', 'Skip', data['skip']]
    dataframe.loc[4] = ['', 'Interrupt', data['interrupt']]
    dataframe.loc[5] = ['', 'Cancel', data['cancel']]
    dataframe.loc[6] = ['', 'Pass', data['pass']]
    dataframe.loc[7] = ['', 'White-Board', data['tests'][0]['whiteboard']]

    # Loop through the 'tests' list in the JSON data and add rows
    for i, test in enumerate(data['tests']):
        dataframe.loc[i + 8] = [test['name'],
                                test['status'], test['fail_reason']]

    # Save the DataFrame to a Excel file
    dataframe.to_excel('Analysis.xlsx', index=False)

    # Save the DataFrame to a HTML page
    summary = {
            'Analysis_Type': 'New',
            'Regression': None,
            'Solved': None,
            'Diff': None,
            'New': {
                'Name': dataframe.loc[0].iat[-1],
                'Fail': dataframe.loc[1].iat[-1],
                'Error': dataframe.loc[2].iat[-1],
                'Skip': dataframe.loc[3].iat[-1],
                'Interrupt': dataframe.loc[4].iat[-1],
                'Cancel': dataframe.loc[5].iat[-1],
                'Pass': dataframe.loc[6].iat[-1],
                'White-Board': dataframe.loc[7].iat[-1]
            }
    }
    json_object = json.dumps(summary, indent=4)
    if "--new-analysis" in sys.argv:
        print(json_object)
        analysis_to_html(summary)


def comparison_analysis(excel, data):
    '''
        This function is used to generate an excel sheet which gives delta comparison
        of two test avocado based test runs. Using excel sheet produced from the
        function: test_analysis(data) and results.json as inputs, it generate a .xlsx
        file as output.
    '''

    # store test_names in existing excel file
    old_dataframe = pd.read_excel(excel)
    test_names = old_dataframe[old_dataframe.columns[0]]
    test_names = [test_names[x] for x in range(8, len(old_dataframe.index))]

    # Extract job name from debuglog attribute
    job_name = data['debuglog'].split('/')[-2]

    # Create a DataFrame
    new_dataframe = pd.DataFrame(columns=['Status', 'Fail Reason'])

    # Add rows for job name and other attributes
    new_dataframe.loc[0] = ['Job Name', job_name]
    new_dataframe.loc[1] = ['Fail', data['failures']]
    new_dataframe.loc[2] = ['Error', data['errors']]
    new_dataframe.loc[3] = ['Skip', data['skip']]
    new_dataframe.loc[4] = ['Interrupt', data['interrupt']]
    new_dataframe.loc[5] = ['Cancel', data['cancel']]
    new_dataframe.loc[6] = ['Pass', data['pass']]
    new_dataframe.loc[7] = ['White-Board', data['tests'][0]['whiteboard']]

    # Loop through the 'tests' list in the JSON data and add rows
    for i, test in enumerate(data['tests']):
        found = 0
        for j in range(len(test_names)):
            if test['name'] == test_names[j]:
                new_dataframe.loc[j + 8] = [test['status'],
                                            test['fail_reason']]
                found = 1
                break
        if found == 0:
            new = [test['name']]
            for i in range(len(old_dataframe.columns)-1):
                new.append("")
            old_dataframe.loc[len(old_dataframe.index)] = new
            new_dataframe.loc[len(old_dataframe.index) -
                              1] = [test['status'], test['fail_reason']]

    # Save the DataFrame to a Excel file
    final_res = pd.concat([old_dataframe, new_dataframe], axis=1)
    final_res.to_excel(excel, index=False)

    # Add the Result column to compare two results
    if "--compare-two-results" in sys.argv:
        regression_count = 0
        solved_count = 0
        difference_count = 0
        dataframe = pd.read_excel(excel)
        results = []
        for i in range(len(dataframe.index)):
            if dataframe.loc[i].iat[-2] == dataframe.loc[i].iat[-4]:
                results.append("")
            else:
                if dataframe.loc[i].iat[-4] == "PASS" and not pd.isnull(dataframe.loc[i].iat[-2]):
                    results.append("REGRESSION")
                    regression_count += 1
                elif dataframe.loc[i].iat[-2] == "PASS" and not pd.isnull(dataframe.loc[i].iat[-4]):
                    results.append("SOLVED")
                    solved_count += 1
                elif pd.isnull(dataframe.loc[i].iat[-4]) or pd.isnull(dataframe.loc[i].iat[-2]):
                    results.append("")
                else:
                    results.append("DIFF")
                    difference_count += 1

        result_dataframe = pd.DataFrame(columns=['Result'])
        for i in range(8, len(results)):
            result_dataframe.loc[i] = results[i]

        # Save the DataFrame to a Excel file
        final_dataframe = pd.concat([dataframe, result_dataframe], axis=1)
        final_dataframe.to_excel(excel, index=False)

        # Save the DataFrame to a HTML page
        summary = {
            'Analysis_Type': 'Comparison',
            'Regression': regression_count,
            'Solved': solved_count,
            'Diff': difference_count,
            'New': {
                'Name': new_dataframe.loc[0].iat[-1],
                'Fail': new_dataframe.loc[1].iat[-1],
                'Error': new_dataframe.loc[2].iat[-1],
                'Skip': new_dataframe.loc[3].iat[-1],
                'Interrupt': new_dataframe.loc[4].iat[-1],
                'Cancel': new_dataframe.loc[5].iat[-1],
                'Pass': new_dataframe.loc[6].iat[-1],
                'White-Board': new_dataframe.loc[7].iat[-1]
            }
        }
        json_object = json.dumps(summary, indent=4)
        print(json_object)
        analysis_to_html(summary)


def deco(excel):
    '''
        This function is used to conditionally format the xlsx file
        Libraries used: ExcelWriter, openpyxl
    '''

    # Create a sample DataFrame
    dataframe = pd.read_excel(excel)

    # Create a Pandas ExcelWriter object and write to Excel file
    excel_writer = pd.ExcelWriter(excel, engine='openpyxl')
    dataframe.to_excel(excel_writer, sheet_name='Sheet1', index=False)

    # Access the workbook and worksheet objects
    workbook = excel_writer.book
    worksheet = excel_writer.sheets['Sheet1']

    # Column Width
    worksheet.column_dimensions['A'].width = 60
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 80
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 80
    worksheet.column_dimensions['F'].width = 20

    # Apply styles to the entire sheet
    for row in worksheet.iter_rows(min_row=2, max_row=len(dataframe) + 1):
        for cell in row:
            cell.font = Font(size=15)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin',
                                            color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = Font(size=18, bold=True)  # White text color
        # Blue background color
        cell.fill = PatternFill(start_color='ADD8E6',
                                end_color='ADD8E6', fill_type='solid')

    # Conditional formatting for the "Result" column if present
    try:
        for idx, value in enumerate(dataframe['Result'], start=2):
            cell = worksheet.cell(row=idx, column=6)
            if value == 'DIFF':
                cell.fill = PatternFill(
                    start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            elif value == 'SOLVED':
                cell.fill = PatternFill(
                    start_color='39E75F', end_color='39E75F', fill_type='solid')  # Green
            elif value == 'REGRESSION':
                cell.fill = PatternFill(
                    start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
    except Exception as e:
        pass

    # Save the styled Excel file
    workbook.save(excel)


def analysis_to_html(summary):
    '''
        This function is used convert the .xlsx output to .html
    '''

    analysis_type = summary['Analysis_Type']

    # Read Excel file
    excel_file = pd.ExcelFile('Analysis.xlsx')
    sheet_names = excel_file.sheet_names

    # Function to apply color based on status
    def apply_formatting(status):
        if status == 'PASS':
            return ('green', 'lightgreen')  # Light green
        elif status == 'FAIL':
            return ('red', 'lightcoral')  # Light red
        elif status == 'ERROR':
            return ('orange', 'lightsalmon')  # Light orange
        else:
            return ('black', 'white')

    # Read each sheet and convert to HTML with custom formatting
    html_tables = {}
    summary_tables = {}
    for sheet_name in sheet_names:
        dataFrame = excel_file.parse(sheet_name)

        # Replace NaN values with blank cells
        dataFrame = dataFrame.fillna('')

        # Apply formatting to "Status" columns
        for col in dataFrame.columns:
            if col.startswith('Status'):
                dataFrame.loc[8:, col] = dataFrame.loc[7:, col].apply(lambda x: f'<span style="color: {apply_formatting(x)[0]}">{x}</span>')

        # Apply formatting to "Result" column
        if 'Result' in dataFrame.columns:
            dataFrame['Result'] = dataFrame['Result'].apply(lambda x: f'<div class="{x.lower()}">{x}</div>')

        # Convert DataFrame to HTML Tables
        if analysis_type == "New":
            summary_table = dataFrame.iloc[:8, 1:].to_html(index=False, header=False)
        else:
            summary_table = dataFrame.iloc[:8, 1:-1].to_html(index=False, header=False)
        summary_table = summary_table.replace('<table border="1" class="dataframe">', '<table border="1" class="table2">')
        summary_tables[sheet_name] = summary_table

        html_table = dataFrame.iloc[8:].to_html(escape=False, index=False, border=1)
        html_table = html_table.replace('<table border="1" class="dataframe">', '<table id="Main_Table" border="1" class="dataframe">')
        html_tables[sheet_name] = html_table

    # Apply styling to the html table
    css_style = """
    <style>
        .dataframe {
            border-collapse: collapse;
            border: 1px solid black;
            font-weight: bold;
        }
        .input-container {
            display: flex;
            justify-content: center;
            gap: 20px;
        }
        h1 {
            text-align: center;
        }
        select {
            font-size: 16px;
        }
        th {
            border: 1px solid black;
            padding: 8px;
            background-color: #add8e6;
            font-size: 20px;
            text-align: center;
        }
        td {
            border: 1px solid black;
            padding: 8px;
        }
        td:has(> div[class="diff"]) {
            color: white;
            background-color: orange;
        }
        td:has(> div[class="solved"]) {
            color: white;
            background-color: green;
        }
        td:has(> div[class="regression"]) {
            color: white;
            background-color: red;
        }
        .table1 {
            border-collapse: collapse;
            border: 1px solid black;
            font-weight: bold;
        }
        .table1 td {
            width: 100px;
            text-align: center;
        }
        .table2 {
            border-collapse: collapse;
            border: 1px solid black;
            font-weight: bold;
        }
        .table2 td {
            width: 100px;
            text-align: center;
            border: 1px solid black;
            background-color: #f0e3ce;
        }
        a {
            text-decoration: none;
            color: #023E8A;
        }
        a:hover {
            color: #00B4D8;
        }
    </style>
    """

    # JavaScript code to handle dropdown menu selection and filter the table
    js_script = """
    <script>
        const resultDropdown = document.getElementById("result-dropdown");
        const statusDropdown = document.getElementById("status-dropdown");
        const statusOneDropdown = document.getElementById("status1-dropdown");
        const rows = Main_Table.getElementsByTagName("tr");

        const resultColumnIndexes = {
            "result-dropdown": 5,
            "status-dropdown": 1,
            "status1-dropdown": 3,
        }

        const resetDropdowns = (id) => {
            id === resultDropdown?.id ? null : resultDropdown.value = "All"
            id === statusDropdown?.id ? null : statusDropdown.value = "All"
            id === statusOneDropdown?.id ? null : statusOneDropdown.value = "All"
        }

        const filterValues = (event) => {
            const columnIndex = resultColumnIndexes[event.target.id]
            const selectedValue = event.target.value

            for (let i = 1; i < rows.length; i++) {
                const cells = rows[i].getElementsByTagName("td");
                const resultCell = cells[columnIndex];
                const resultText = resultCell.textContent || resultCell.innerText;

                if (selectedValue === "All" || resultText === selectedValue) {
                    rows[i].style.display = "";
                } else {
                    rows[i].style.display = "none";
                }
            }
            resetDropdowns(event.target.id)
        }

        resultDropdown?.addEventListener("input", filterValues);
        statusOneDropdown?.addEventListener("input", filterValues);
        statusDropdown?.addEventListener("input", filterValues);
    </script>
    """

    # Dropdown menu HTML code
    dropdown_result = """
    <div>
        <label for="result-dropdown">Result:</label>
        <select id="result-dropdown">
            <option value="All">All</option>
            <option value="REGRESSION">REGRESSION</option>
            <option value="SOLVED">SOLVED</option>
            <option value="DIFF">DIFF</option>
        </select>
    </div>
    """

    dropdown_status = """
    <div>
        <label for="status-dropdown">Status:</label>
        <select id="status-dropdown">
            <option value="All">All</option>
            <option value="PASS">PASS</option>
            <option value="FAIL">FAIL</option>
            <option value="ERROR">ERROR</option>
            <option value="SKIP">SKIP</option>
            <option value="CANCEL">CANCEL</option>
            <option value="INTERRUPT">INTERRUPT</option>
        </select>
    </div>
    """

    dropdown_status1 = """
    <div>
        <label for="status1-dropdown">Status.1:</label>
        <select id="status1-dropdown">
            <option value="All">All</option>
            <option value="PASS">PASS</option>
            <option value="FAIL">FAIL</option>
            <option value="ERROR">ERROR</option>
            <option value="SKIP">SKIP</option>
            <option value="CANCEL">CANCEL</option>
            <option value="INTERRUPT">INTERRUPT</option>
        </select>
    </div>
    """

    quick_summary = f"""
        <div class="input-container">
            <table class="table1">
                <tr>
                    <td><div class="regression">REGRESSION</div></td>
                    <td><div class="solved">SOLVED</div></td>
                    <td><div class="diff">DIFF</div></td>
                </tr>
                <tr>
                    <td>{summary["Regression"]}</td>
                    <td>{summary["Solved"]}</td>
                    <td>{summary["Diff"]}</td>
                </tr>
            </table>
        </div>
    """

    # Save combined HTML table with JavaScript to 'Analysis.html'
    with open('Analysis.html', 'w') as file:
        file.write(css_style)
        file.write('<h1>Analysis</h1>')

        if analysis_type == "Comparison":
            file.write(quick_summary)

        file.write('<br><div class="input-container">')
        file.write(dropdown_status)
        if analysis_type == "Comparison":
            file.write(dropdown_status1)
            file.write(dropdown_result)
        file.write('</div><br>')

        file.write('<br><div class="input-container">')
        for sheet_name, summary_table in summary_tables.items():
            file.write(summary_table)
        file.write('</div><br>')

        file.write('<br><div class="input-container">')
        for sheet_name, html_table in html_tables.items():
            file.write(html_table)
        file.write(js_script)
        file.write('</div><br>')


def main():
    try:
        if "--new-analysis" in sys.argv:
            with open(sys.argv[-1], 'r') as json_file:
                data = json.load(json_file)
            test_analysis(data)
            deco("Analysis.xlsx")

        elif "--compare-two-results" in sys.argv:
            with open(sys.argv[-2], 'r') as json_file:
                data = json.load(json_file)
            test_analysis(data)
            deco("Analysis.xlsx")

            with open(sys.argv[-1], 'r') as json_file:
                data = json.load(json_file)
            comparison_analysis("Analysis.xlsx", data)
            deco("Analysis.xlsx")

        else:
            raise Exception

    except Exception as e:
        print(str(e))
        print("\nPay attention on the usage:\n"+usage())
        sys.exit(1)


def usage():
    return ("python3 analysis.py --new-analysis <json_file>\n\
python3 analysis.py --compare-two-results <old_json_file> <new_json_file>\n")


if __name__ == '__main__':
    main()
